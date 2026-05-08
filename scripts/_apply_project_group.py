#!/usr/bin/env python3
"""One-off: Add 'Project Group' column to Bdb-317.xlsx and rename one Project ID.

- Appends column 64 named 'Project Group'.
- Sets value to 'CPER' on rows whose Project ID matches CPER_MEMBERS.
- Renames '2023 Spring CPER Chemistry' -> 'Spring Chemistry'.
- Idempotent: re-running won't duplicate the column or re-rename.
"""

from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = REPO_ROOT / "Bdb-317.xlsx"

# Note: '2023 Spring CPER Chemistry' is renamed FIRST to 'Spring Chemistry',
# then 'Spring Chemistry' (which now includes both) gets the CPER tag.
RENAME_MAP = {
    "2023 Spring CPER Chemistry": "Spring Chemistry",
}

CPER_MEMBERS = {
    "Fall Plant Circle",
    "Spring Plant Circle",
    "Fall Plants & Soil",
    "Spring SASS/Polycarbonate Top/Bottom",
    "Spring Sass/VIVAS",
    "Spring Plants & Soil",
    "Spring Chemistry",
    "2022 Fall CPER",
    "2022 Fall CPER FC",
}


def main() -> None:
    print(f"Loading {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Sheet1"]

    headers = [cell.value for cell in ws[1]]
    print(f"  rows={ws.max_row}, cols={ws.max_column}")

    # Locate Project ID column (1-indexed)
    if "Project ID" not in headers:
        raise SystemExit("ERROR: 'Project ID' column not found")
    project_id_col = headers.index("Project ID") + 1
    print(f"  Project ID is column {project_id_col}")

    # Either find existing Project Group column or append new one
    if "Project Group" in headers:
        project_group_col = headers.index("Project Group") + 1
        print(f"  Project Group already exists at col {project_group_col} — will overwrite values")
    else:
        project_group_col = ws.max_column + 1
        ws.cell(row=1, column=project_group_col, value="Project Group")
        print(f"  Added 'Project Group' header at col {project_group_col}")

    # Pass 1: rename
    rename_count = 0
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=project_id_col)
        if cell.value in RENAME_MAP:
            cell.value = RENAME_MAP[cell.value]
            rename_count += 1
    print(f"  Renamed {rename_count} Project ID cells")

    # Pass 2: tag CPER members
    tag_count = 0
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(row=r, column=project_id_col).value
        if pid in CPER_MEMBERS:
            ws.cell(row=r, column=project_group_col, value="CPER")
            tag_count += 1
    print(f"  Tagged {tag_count} rows with Project Group='CPER'")

    print(f"Saving {XLSX_PATH}...")
    wb.save(XLSX_PATH)
    print("Done.")


if __name__ == "__main__":
    main()
